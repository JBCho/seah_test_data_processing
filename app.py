import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border
from copy import copy
import io

# --- 설정 부분 ---
# 사용자는 이 부분에서 파일 이름만 실제 파일에 맞게 수정하면 됩니다.
FILENAME_CONFIG = {
    "template": "시험결과 통합 양식.xlsx",
    "component": "API-X56L2-D 성분시험결과.xlsx",
    "tensile": "API-X56L2-D 인장시험결과.xlsx",
    "impact": "API-X56L2-D 충격시험결과.xlsx",
    "output": "통합_시험_결과_완성본.xlsx"
}

# [수정] 템플릿의 헤더가 시작되는 행 번호
TEMPLATE_HEADER_ROW = 3

# [신규] 템플릿의 열 순서에 맞춘 최종 DataFrame의 열 목록
# 이 목록은 템플릿의 헤더 순서와 정확히 일치해야 합니다.
TEMPLATE_ORDERED_COLS = [
    '시편배치', '생산오더', '제품배치', '제품기호', '외경', '두께', 'Heat No.', '원재료기호', '원재료업체',
    # 성분 1 (19개)
    'C_1', 'Si_1', 'Mn_1', 'P_1', 'S_1', 'Cu_1', 'Ni_1', 'Cr_1', 'Mo_1', 'V_1', 'Nb_1', 'Ti_1', 'Alsol_1', 'Aloxy_1', 'Al_1', 'Ca_1', 'B_1', 'PCM_1', 'CEQ_1',
    # 성분 2 (19개)
    'C_2', 'Si_2', 'Mn_2', 'P_2', 'S_2', 'Cu_2', 'Ni_2', 'Cr_2', 'Mo_2', 'V_2', 'Nb_2', 'Ti_2', 'Alsol_2', 'Aloxy_2', 'Al_2', 'Ca_2', 'B_2', 'PCM_2', 'CEQ_2',
    # 인장 L (4개)
    'Stripe 모재 L방향_YS2 STRESS', 'Stripe 모재 L방향_TS STRESS', 'Stripe 모재 L방향_연신율 EL(%)', 'Stripe 모재 L방향_YR(%)',
    # 인장 T (4개)
    'Stripe 모재 T방향_YS2 STRESS', 'Stripe 모재 T방향_TS STRESS', 'Stripe 모재 T방향_연신율 EL(%)', 'Stripe 모재 T방향_YR(%)',
    # 인장 용접 (4개)
    'Stripe 용접_YS2 STRESS', 'Stripe 용접_TS STRESS', 'Stripe 용접_연신율 EL(%)', 'Stripe 용접_YR(%)',
    # 충격 Base (5개)
    'Base (Transeverse)_온도', 'Base (Transeverse)_1', 'Base (Transeverse)_2', 'Base (Transeverse)_3', 'Base (Transeverse)_Avg',
    # 충격 Weld (5개)
    'Weld Line_온도', 'Weld Line_1', 'Weld Line_2', 'Weld Line_3', 'Weld Line_Avg',
    # 충격 HAZ (5개)
    'HAZ_온도', 'HAZ_1', 'HAZ_2', 'HAZ_3', 'HAZ_Avg'
]


# --- 데이터 처리 함수들 ---

def get_data(filename, sheet_name=0):
    """엑셀 파일을 안전하게 읽어오는 함수"""
    try:
        return pd.read_excel(filename, sheet_name=sheet_name)
    except FileNotFoundError:
        st.error(f"오류: '{filename}' 파일을 찾을 수 없습니다. 스크립트와 같은 폴더에 파일이 있는지 확인하세요.")
        return None
    except Exception as e:
        st.error(f"오류: '{filename}' 파일을 읽는 중 문제가 발생했습니다: {e}")
        return None
    
def get_impact_data_with_multiheader(filename):
    """[수정] 2줄 헤더를 가진 충격 시험 엑셀 파일을 읽고 컬럼명을 정리하는 함수"""
    try:
        df = pd.read_excel(filename, header=[0, 1])
        
        new_columns = []
        for col in df.columns:
            level1 = str(col[0]) if 'Unnamed:' not in str(col[0]) else ''
            level2 = str(col[1]) if 'Unnamed:' not in str(col[1]) else ''
            
            if level1 and level2:
                new_columns.append(f"{level1}_{level2}")
            elif level1:
                new_columns.append(level1)
            else:
                new_columns.append(level2)
        
        df.columns = new_columns
        
        # [수정] 원본 컬럼명(다중 헤더)도 유지하여 키 컬럼 접근에 사용
        # 예: ('시편배치', '시편배치') -> '시편배치'
        # df.columns에서 '시편배치', '외경', '두께', 'Heat No.'를 포함하는 컬럼을 찾아 단일 이름으로 매핑
        # 이 부분은 get_impact_data_with_multiheader가 단일 이름으로 잘 변환한다고 가정하고,
        # process_impact_data에서 처리하도록 수정합니다.
        
        return df

    except FileNotFoundError:
        st.error(f"오류: '{filename}' 파일을 찾을 수 없습니다.")
        return None
    except Exception as e:
        st.error(f"오류: '{filename}' 파일을 읽는 중 문제가 발생했습니다: {e}")
        return None

def process_component_data(df):
    """[수정] 규칙 2: 성분 시험 데이터 처리 (복합 키 사용)"""
    if df is None: return pd.DataFrame()
    
    # [수정] 복합 키로 사용할 컬럼 정의
    base_key_cols = ['시편배치', '외경', '두께', 'Heat No.']
    
    # [수정] 키 컬럼이 모두 존재하는지 확인
    if not all(col in df.columns for col in base_key_cols):
        st.error(f"성분 시험 파일에 필수 키 컬럼({base_key_cols}) 중 일부가 없습니다.")
        return pd.DataFrame()

    # [신규] '시편배치'의 앞 8자리를 키로 사용
    df['시편배치_키'] = df['시편배치'].str[:8]
    key_cols = ['시편배치_키', '외경', '두께', 'Heat No.']

    # 필요한 성분 컬럼 목록 (템플릿 기준)
    comp_cols = ['C', 'Si', 'Mn', 'P', 'S', 'Cu', 'Ni', 'Cr', 'Mo', 'V', 'Nb', 'Ti', 'Alsol', 'Aloxy', 'Al', 'Ca', 'B', 'PCM', 'CEQ']
    
    # [수정] 기본 정보 컬럼 (키 컬럼 제외, 원본 '시편배치' 추가)
    info_cols = ['생산오더', '제품배치', '제품기호', '원재료기호', '원재료업체', '시편배치']
    
    processed_data = {}

    # [수정] 복합 키로 그룹화
    for key, group in df.groupby(key_cols):
        last_two = group.tail(2)
        
        # 1. 기본 정보 추출 (첫 번째 행에서만)
        # info_cols에 없는 컬럼이 있을 수 있으므로 .get() 사용
        info_data = {col: last_two.iloc[0].get(col) for col in info_cols}

        # [신규] '시편배치' 값을 8자리 키 값으로 덮어쓰기 (v2.2)
        if '시편배치' in info_data:
            info_data['시편배치'] = last_two.iloc[0].get('시편배치_키')

        # 2. 성분 데이터 추출 및 컬럼명 변경
        row_data = {}
        for i, (idx, row) in enumerate(last_two.iterrows()):
            suffix = f'_{i+1}' # _1, _2
            for col in comp_cols:
                if col in row:
                    row_data[col + suffix] = row[col]
        
        processed_data[key] = {**info_data, **row_data}
        
    result_df = pd.DataFrame.from_dict(processed_data, orient='index')
    # [수정] 인덱스 이름 설정
    result_df.index.names = key_cols
    return result_df


def process_tensile_data(df):
    """[수정] 규칙 3: 인장 시험 데이터 처리 (복합 키 사용)"""
    if df is None: return pd.DataFrame()

    # [수정] 복합 키로 사용할 컬럼 정의
    base_key_cols = ['시편배치', '외경', '두께', 'Heat No.']

    # [수정] 키 컬럼이 모두 존재하는지 확인
    if not all(col in df.columns for col in base_key_cols):
        st.error(f"인장 시험 파일에 필수 키 컬럼({base_key_cols}) 중 일부가 없습니다.")
        return pd.DataFrame()
    
    # [신규] '시편배치'의 앞 8자리를 키로 사용
    df['시편배치_키'] = df['시편배치'].str[:8]
    key_cols = ['시편배치_키', '외경', '두께', 'Heat No.']
    
    # 처리할 방향과 결과 컬럼 정의
    directions = ["Stripe 모재 L방향", "Stripe 모재 T방향", "Stripe 용접"]
    result_cols = ["YS2 STRESS", "TS STRESS", "연신율 EL(%)", "YR(%)"]
    
    all_data = {}

    # [수정] 복합 키로 그룹화
    for key, group in df.groupby(key_cols):
        key_data = {}
        for direction in directions:
            dir_group = group[group['시편 위치/방향'] == direction]
            if not dir_group.empty:
                last_test = dir_group.iloc[-1]
                for col in result_cols:
                    # [수정] 컬럼이 없을 경우 None 반환
                    key_data[f"{direction}_{col}"] = last_test.get(col, None)
            else:
                for col in result_cols:
                    key_data[f"{direction}_{col}"] = None
        all_data[key] = key_data
        
    result_df = pd.DataFrame.from_dict(all_data, orient='index')
    # [수정] 인덱스 이름 설정
    result_df.index.names = key_cols
    return result_df


def process_impact_data(df):
    """[수정] 규칙 4: 충격 시험 데이터 처리 (복합 키 사용)"""
    if df is None: return pd.DataFrame()

    # [수정] 동적으로 컬럼명 찾기 (정리된 컬럼명 기준)
    # get_impact_data_with_multiheader 함수가 '시편배치_시편배치' -> '시편배치' 등으로
    # 잘 정리해준다고 가정합니다.
    def find_col(df, keyword):
        # 먼저 정확히 일치하는 이름 찾기
        if keyword in df.columns:
            return keyword
        # 없다면 키워드를 포함하는 컬럼 찾기
        for col in df.columns:
            if keyword in col:
                return col
        return None

    specimen_col = find_col(df, '시편배치')
    od_col = find_col(df, '외경')
    thick_col = find_col(df, '두께')
    heat_col = find_col(df, 'Heat No.')
    notch_col = find_col(df, 'Notch 위치')
    
    # [수정] 복합 키 컬럼 리스트
    base_key_cols_found = [specimen_col, od_col, thick_col, heat_col]
    
    if not all(base_key_cols_found + [notch_col]):
        st.error(f"충격 시험 파일에서 필수 키/Notch 컬럼을 찾을 수 없습니다. (찾은 컬럼: {base_key_cols_found}, {notch_col})")
        return pd.DataFrame()

    # [신규] '시편배치'의 앞 8자리를 키로 사용
    df['시편배치_키'] = df[specimen_col].str[:8]
    key_cols = ['시편배치_키', od_col, thick_col, heat_col] # [수정] specimen_col 대신 '시편배치_키' 사용
    # [수정] 인덱스 이름도 통일
    key_col_names = ['시편배치_키', '외경', '두께', 'Heat No.']


    # 컬럼 접두사 정의
    temp_col_prefix = '온도(˚C)'
    energy_col_prefix = '에너지(J) SIZE 10보정'

    locations = ["Base (Transeverse)", "Weld Line", "HAZ"]
    all_data = {}
    
    # [수정] 복합 키로 그룹화
    for key, group in df.groupby(key_cols):
        key_data = {}
        for loc in locations:
            loc_group = group[group[notch_col] == loc]
            
            if not loc_group.empty:
                last_test_row = loc_group.iloc[-1]
                
                temp_values = []
                for i in range(1, 7):
                    col_name = f'{temp_col_prefix}_{i}'
                    if col_name in last_test_row and pd.notna(last_test_row[col_name]):
                        temp_values.append(last_test_row[col_name])
                
                test_temperature = None
                if temp_values:
                    mode_series = pd.Series(temp_values).mode()
                    if not mode_series.empty:
                        test_temperature = mode_series.iloc[0]

                val1 = last_test_row.get(f'{energy_col_prefix}_1', None)
                val2 = last_test_row.get(f'{energy_col_prefix}_2', None)
                val3 = last_test_row.get(f'{energy_col_prefix}_3', None)
                
                valid_values = [v for v in [val1, val2, val3] if pd.notna(v) and isinstance(v, (int, float))]
                
                key_data[f'{loc}_온도'] = test_temperature
                key_data[f'{loc}_1'] = val1
                key_data[f'{loc}_2'] = val2
                key_data[f'{loc}_3'] = val3
                key_data[f'{loc}_Avg'] = sum(valid_values) / len(valid_values) if valid_values else None
            else:
                for col in ['온도', '1', '2', '3', 'Avg']:
                    key_data[f'{loc}_{col}'] = None
        all_data[key] = key_data
    
    result_df = pd.DataFrame.from_dict(all_data, orient='index')
    # [수정] 인덱스 이름 설정
    result_df.index.names = key_col_names
    return result_df


def reorder_final_dataframe(final_df, template_cols):
    """
    [신규] 병합된 DataFrame을 템플릿 순서에 맞게 재정렬하고
    누락된 컬럼은 None으로 채우는 함수
    """
    final_df_ordered = pd.DataFrame()
    for col in template_cols:
        if col in final_df.columns:
            final_df_ordered[col] = final_df[col]
        else:
            # 템플릿에 필요한 컬럼이 병합된 데이터에 없으면 빈 컬럼 추가
            final_df_ordered[col] = None 
    return final_df_ordered


def write_data_to_excel(wb, final_df_ordered):
    """
    [신규] 준비된 DataFrame을 템플릿 엑셀 워크북에
    서식을 복사하며 쓰는 함수
    """
    try:
        ws = wb.active
    except Exception as e:
        st.error(f"엑셀 워크북에서 활성 시트를 찾는 중 오류 발생: {e}")
        return None

    # 데이터 쓰기 시작할 행 (기존 데이터 다음 행)
    start_row = ws.max_row + 1
    # 서식을 복사할 템플릿 행 (기존 데이터의 마지막 행)
    style_template_row = ws.max_row if ws.max_row >= TEMPLATE_HEADER_ROW else TEMPLATE_HEADER_ROW
    
    # [수정] 템플릿 헤더의 총 컬럼 수 (서식 복사 기준)
    # TEMPLATE_ORDERED_COLS 리스트의 길이를 사용
    total_template_cols = len(TEMPLATE_ORDERED_COLS)

    for index, row_data in final_df_ordered.iterrows():
        current_row = start_row + index
        
        # [수정] 순서가 보장된 final_df_ordered의 값을 순서대로 입력
        for col_idx, value in enumerate(row_data.values, 1):
            if pd.isna(value):
                value = None
            ws.cell(row=current_row, column=col_idx, value=value)

        # 서식 복사
        for col_num in range(1, total_template_cols + 1):
            template_cell = ws.cell(row=style_template_row, column=col_num)
            if not template_cell:
                continue
                
            new_cell = ws.cell(row=current_row, column=col_num)
            
            if template_cell.has_style:
                new_cell.font = copy(template_cell.font)
                new_cell.border = copy(template_cell.border)
                new_cell.fill = copy(template_cell.fill)
                new_cell.number_format = copy(template_cell.number_format)
                new_cell.protection = copy(template_cell.protection)
                new_cell.alignment = copy(template_cell.alignment)
    
    return wb


def main():
    """메인 실행 함수 (로컬 실행용)"""
    print("--- 데이터 통합 작업을 시작합니다 ---")

    df_comp_raw = get_data(FILENAME_CONFIG['component'])
    df_tens_raw = get_data(FILENAME_CONFIG['tensile'])
    df_impa_raw = get_impact_data_with_multiheader(FILENAME_CONFIG['impact'])

    if any(df is None for df in [df_comp_raw, df_tens_raw, df_impa_raw]):
        print("필수 데이터 파일이 없어 작업을 중단합니다.")
        return

    print("1/4: 성분, 인장, 충격 데이터 처리 중...")
    processed_comp = process_component_data(df_comp_raw)
    processed_tens = process_tensile_data(df_tens_raw)
    processed_impa = process_impact_data(df_impa_raw)

    print("2/4: 처리된 데이터 병합 중...")
    final_df = processed_comp.join(processed_tens, how='outer')
    final_df = final_df.join(processed_impa, how='outer')

    # [수정] 인덱스(복합 키)를 컬럼으로 변환
    final_df.reset_index(inplace=True)
    # [삭제] 'index' 컬럼 이름 변경 로직 (reset_index가 자동으로 인덱스 이름 사용)

    print("3/4: 엑셀 템플릿 파일에 데이터 쓰는 중...")
    try:
        wb = openpyxl.load_workbook(FILENAME_CONFIG['template'])
    except FileNotFoundError:
        print(f"오류: 템플릿 파일 '{FILENAME_CONFIG['template']}'을 찾을 수 없습니다.")
        return
    except Exception as e:
        print(f"템플릿 파일 로드 중 오류: {e}")
        return

    # [신규] DataFrame을 템플릿 순서로 재정렬
    final_df_ordered = reorder_final_dataframe(final_df, TEMPLATE_ORDERED_COLS)

    # [신규] 엑셀 쓰기 함수 호출
    wb = write_data_to_excel(wb, final_df_ordered)

    if wb is None:
        print("엑셀 파일 쓰기에 실패했습니다.")
        return

    print(f"4/4: '{FILENAME_CONFIG['output']}' 파일 저장 중...")
    try:
        wb.save(FILENAME_CONFIG['output'])
        print(f"--- 작업 완료! 결과가 '{FILENAME_CONFIG['output']}' 파일에 저장되었습니다. ---")
    except PermissionError:
        print(f"오류: '{FILENAME_CONFIG['output']}' 파일이 다른 프로그램에서 열려있어 저장할 수 없습니다. 파일을 닫고 다시 시도해주세요.")
    except Exception as e:
        print(f"파일 저장 중 오류가 발생했습니다: {e}")


# --- Streamlit 페이지 설정 ---
st.set_page_config(page_title="시험 결과 통합 자동화 툴", layout="wide")

st.title("🔬 시험 결과 통합 자동화 툴")
st.write("아래 4개의 엑셀 파일을 업로드한 후 버튼을 누르면, 규칙에 따라 데이터를 통합하고 서식을 유지한 최종 결과 파일을 다운로드할 수 있습니다.")
st.write("**[v2.2]** '시편배치'(앞 8자리), '외경', '두께', 'Heat No.'를 기준으로 데이터를 통합하고, 결과 파일에도 8자리 '시편배치'를 표시합니다.")


# --- UI 부분 ---
st.subheader("1. 파일 업로드")
col1, col2 = st.columns(2)
with col1:
    template_file = st.file_uploader("📂 **양식 파일** (.xlsx)", type=['xlsx'])
    component_file = st.file_uploader("📂 **성분 시험 결과** (.xlsx)", type=['xlsx'])
with col2:
    tensile_file = st.file_uploader("📂 **인장 시험 결과** (.xlsx)", type=['xlsx'])
    impact_file = st.file_uploader("📂 **충격 시험 결과** (.xlsx)", type=['xlsx'])

st.divider()

if all([template_file, component_file, tensile_file, impact_file]):
    st.subheader("2. 결과 생성")
    if st.button("🚀 결과 생성 및 다운로드", type="primary", use_container_width=True):
        with st.spinner('데이터를 처리하고 엑셀 파일을 생성하는 중입니다... 잠시만 기다려주세요.'):
            # 1. 파일 읽기 (UploadedFile 객체 전달)
            df_comp_raw = get_data(component_file)
            df_tens_raw = get_data(tensile_file)
            df_impa_raw = get_impact_data_with_multiheader(impact_file)

            # 2. 데이터 처리
            processed_comp = process_component_data(df_comp_raw)
            processed_tens = process_tensile_data(df_tens_raw)
            processed_impa = process_impact_data(df_impa_raw)
            
            # [수정] 처리 중 오류가 발생했는지 확인
            if any(df.empty for df in [processed_comp, processed_tens, processed_impa] if df is not None):
                st.error("데이터 처리 중 오류가 발생했습니다. 각 파일에 필요한 키 컬럼이 모두 있는지 확인해주세요.")
                st.stop()

            # 3. 데이터 병합
            final_df = processed_comp.join(processed_tens, how='outer')
            final_df = final_df.join(processed_impa, how='outer')
            
            # [수정] 인덱스(복합 키)를 컬럼으로 변환
            final_df.reset_index(inplace=True)
            # [삭제] 'index' 컬럼 이름 변경 로직

            # 4. 템플릿에 데이터 쓰기
            try:
                wb = openpyxl.load_workbook(template_file)
            except Exception as e:
                st.error(f"템플릿 파일을 여는 중 오류가 발생했습니다: {e}")
                st.stop()

            # [신규] DataFrame을 템플릿 순서로 재정렬
            final_df_ordered = reorder_final_dataframe(final_df, TEMPLATE_ORDERED_COLS)

            # [신규] 엑셀 쓰기 함수 호출
            wb = write_data_to_excel(wb, final_df_ordered)

            if wb is None:
                st.error("엑셀 파일 쓰기에 실패했습니다.")
                st.stop()

            # 5. 최종 엑셀 파일을 메모리에 저장
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)

            st.session_state.download_ready = True
            st.session_state.output_buffer = output_buffer
            st.session_state.error_occurred = False

        if not st.session_state.get('error_occurred', False):
            st.success("✅ 처리가 완료되었습니다! 아래 버튼을 눌러 파일을 다운로드하세요.")
    
    if 'download_ready' in st.session_state and st.session_state.download_ready:
        st.download_button(
            label="📥 '통합_시험_결과_완성본.xlsx' 다운로드",
            data=st.session_state.output_buffer,
            file_name="통합_시험_결과_완성본.xlsx",
            mime="application/vnd.ms-excel",
            use_container_width=True
        )

else:
    st.info("💡 4개의 파일을 모두 업로드하면 결과 생성 버튼이 나타납니다.")

# [신규] 로컬 실행을 위한 엔트리 포인트
if __name__ == "__main__":
    # Streamlit이 실행 중인지 확인
    try:
        st.runtime.get_instance()
    except RuntimeError:
        # Streamlit이 실행 중이 아니면 main() 함수 호출
        main()


